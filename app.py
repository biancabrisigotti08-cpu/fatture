"""
Estrattore Fatture Web - Flask Backend
"""
import zipfile
import io
import re
import os
import xml.etree.ElementTree as ET
from pathlib import Path
from flask import Flask, request, send_file, jsonify, render_template_string
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pypdf import PdfReader
app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024  # 500MB max upload
HEADER_BG   = "FF6B35"
HEADER_FONT = "FFFFFF"
IMPORTO_BOLLO = 2.00

# ─── Parser XML FatturaPA ──────────────────────────────────────────────────────
def find_text(element, *tags):
   for tag in tags:
       found = element.find('.//' + tag)
       if found is not None and found.text:
           return found.text.strip()
       for child in element.iter():
           local = child.tag.split('}')[-1] if '}' in child.tag else child.tag
           if local == tag and child.text:
               return child.text.strip()
   return ""

def parse_xml_fattura(xml_bytes):
   try:
       root = ET.fromstring(xml_bytes)
   except ET.ParseError as e:
       raise ValueError(f"XML non valido: {e}")
   cedente_block = None
   cessionario_block = None
   for child in root.iter():
       local = child.tag.split('}')[-1] if '}' in child.tag else child.tag
       if local == "CedentePrestatore" and cedente_block is None:
           cedente_block = child
       if local == "CessionarioCommittente" and cessionario_block is None:
           cessionario_block = child
   cedente     = find_text(cedente_block, "Denominazione")     if cedente_block     else ""
   cessionario = find_text(cessionario_block, "Denominazione") if cessionario_block else ""
   numero_documento = find_text(root, "Numero")
   data_documento   = find_text(root, "Data")
   righe = []
   for linea in root.iter():
       local = linea.tag.split('}')[-1] if '}' in linea.tag else linea.tag
       if local != "DettaglioLinee":
           continue
       descrizione_raw = find_text(linea, "Descrizione")
       prezzo_totale   = find_text(linea, "PrezzoTotale")
       try:
           if abs(float(prezzo_totale)) == IMPORTO_BOLLO:
               continue
       except (ValueError, TypeError):
           pass
       telaio = ""
       descrizione = ""
       if descrizione_raw:
           m = re.match(r'^(\S+)\s+RMK\S+\s+(.+)$', descrizione_raw.strip(), re.IGNORECASE)
           if m:
               telaio      = m.group(1).strip()
               descrizione = m.group(2).strip()
           else:
               parts = descrizione_raw.strip().split(None, 1)
               telaio      = parts[0] if parts else ""
               descrizione = parts[1] if len(parts) > 1 else ""
       righe.append({
           "telaio": telaio,
           "descrizione": descrizione,
           "prezzo_totale": prezzo_totale,
       })
   return {
       "cedente": cedente,
       "cessionario": cessionario,
       "numero_documento": numero_documento,
       "data_documento": data_documento,
       "righe": righe,
   }

# ─── Parser PDF — formato generico (PSA + Romana Diesel + altri) ──────────────
def parse_pdf_fattura(pdf_bytes):
   text = ""
   reader = PdfReader(io.BytesIO(pdf_bytes))
   for page in reader.pages:
       t = page.extract_text()
       if t:
           text += t + "\n"
   lines = text.splitlines()
   # Cedente
   cedente = ""
   for i, l in enumerate(lines):
       if re.search(r'cedente|prestatore|fornitore', l, re.IGNORECASE):
           for j in range(i+1, min(i+6, len(lines))):
               nl = lines[j].strip()
               if nl and not re.match(r'^[A-Z\s/()]{6,}$', nl):
                   cedente = nl
                   break
           if cedente:
               break
   if not cedente:
       for l in lines[:10]:
           nl = l.strip()
           if nl and len(nl) > 3:
               cedente = nl
               break
   # Cessionario
   cessionario = ""
   for i, l in enumerate(lines):
       if re.search(r'cessionario|committente|spett', l, re.IGNORECASE):
           for j in range(i+1, min(i+6, len(lines))):
               nl = lines[j].strip()
               if nl and not re.match(r'^[A-Z\s/()]{6,}$', nl):
                   cessionario = nl
                   break
           if cessionario:
               break
   # Numero documento
   numero_documento = ""
   # Formato VW: "TD01 (fattura) 000866019 16980504532 30-03-2026"
   m = re.search(r'TD0\d\s*\([^)]+\)\s*(\d{6,})', text, re.IGNORECASE)
   if m:
       numero_documento = m.group(1).strip()
   else:
       # Formato VW alternativo: "NUMERO DOCUMENTO ART. 73 NUMERO DOCUMENTO\n...000866019"
       m = re.search(r'NUMERO\s+DOCUMENTO(?:\s+ART[.\s]+\d+)?\s*[\n\r]+\s*(?:ART[.\s]+\d+\s*[\n\r]+\s*)?(\d{6,})', text, re.IGNORECASE)
       if m:
           numero_documento = m.group(1).strip()
   if not numero_documento:
       # Formato PSA: "NUMERO DOCUMENTO\n1181358498"
       m = re.search(r'NUMERO\s+DOCUMENTO\s*[\n\r]+\s*(\S+)', text, re.IGNORECASE)
       if m:
           val = m.group(1).strip()
           if not re.match(r'^[A-Z]+$', val, re.IGNORECASE):
               numero_documento = val
   if not numero_documento:
       # Formato Romana Diesel: "Numero\nG000617"
       m = re.search(r'\bNumero\b\s*[\n\r]+\s*([A-Z0-9]+)', text, re.IGNORECASE)
       if m:
           numero_documento = m.group(1).strip()
   if not numero_documento:
       m = re.search(r'\bNumero\b\s+([A-Z0-9]{4,})', text, re.IGNORECASE)
       if m:
           numero_documento = m.group(1).strip()
   # Data documento
   data_documento = ""
   m = re.search(r'DATA\s+DOCUMENTO\s*[\n\r\s]+(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})', text, re.IGNORECASE)
   if m:
       data_documento = m.group(1).strip()
   else:
       m = re.search(r'\bdata\b\s*[\n\r]+\s*(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})', text, re.IGNORECASE)
       if m:
           data_documento = m.group(1).strip()
       else:
           m = re.search(r'(\d{1,2}[-/]\d{2}[-/]\d{4})', text)
           if m:
               data_documento = m.group(1).strip()
   is_psa_format     = bool(re.search(r'RMK\w+', text))
   is_romana_format  = bool(re.search(r'RIF\.TARGA', text, re.IGNORECASE))
   is_vw_format      = bool(re.search(r'Tipo dato:TELAIO', text, re.IGNORECASE))
   righe = []
   if is_vw_format:
       # ── Parser Volkswagen — usa numeri di riga come separatori ──
       # Il testo reale dai log mostra che prima di ogni "ADDEBITO PENALE PER"
       # c'è un numero di riga (es. "30\nADDEBITO PENALE PER DANNI -\n...")
       # Usiamo il pattern "\nN.\n" o "\nN \n" come separatore di blocco.
       # Questo funziona anche quando il blocco va a capo pagina perché
       # il numero di riga appare sempre prima della descrizione.
       import pdfplumber
       # Estrai testo pagina per pagina con pdfplumber (più preciso sui \n)
       full_pages_text = []
       try:
           with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
               for page in pdf.pages:
                   t = page.extract_text(layout=True)
                   if t:
                       full_pages_text.append(t)
       except Exception:
           full_pages_text = [text]
       # Unisci tutto il testo delle pagine con un marcatore di pagina
       combined = '\n'.join(full_pages_text)
       # Cerca tutti i numeri di riga nel testo
       # Pattern: numero intero su riga propria (o con spazi) seguito da ADDEBITO
       # Dal PDF reale: "30 ADDEBITO PENALE PER DANNI -"
       # oppure su righe separate: "30\nADDEBITO PENALE PER DANNI"
       row_pattern = re.compile(
           r'(\d+)\s+ADDEBITO\s+PENALE\s+PER',
           re.IGNORECASE
       )
       matches = list(row_pattern.finditer(combined))
       print(f"=== Blocchi VW trovati: {len(matches)} ===")
       for idx, match in enumerate(matches):
           block_start = match.start()
           block_end = matches[idx + 1].start() if idx + 1 < len(matches) else len(combined)
           blocco_text = combined[block_start:block_end]
           # Descrizione
           desc_upper = blocco_text.upper()
           if 'ELEMENTI TECNICI' in desc_upper:
               desc_val = 'Addebito Penale per Elementi Tecnici Mancanti'
           elif 'ECCEDENZA' in desc_upper or 'CHILOMETRICA' in desc_upper:
               desc_val = 'Addebito Penale per Eccedenza Chilometrica'
           elif 'DANNI' in desc_upper:
               desc_val = 'Addebito Penale per Danni'
           else:
               desc_val = 'Addebito Penale'
           # Telaio — cerca in tutto il blocco
           m_telaio = re.search(
               r'Tipo\s*dato:\s*TELAIO\s*[\n\r]+\s*Rif\.\s*testo:\s*(\S+)',
               blocco_text, re.IGNORECASE
           )
           if not m_telaio:
               m_telaio = re.search(r'Rif\.\s*testo:\s*([A-Z0-9]{17})\b', blocco_text, re.IGNORECASE)
           telaio_val = m_telaio.group(1).strip() if m_telaio else ""
           # Targa — cerca in tutto il blocco
           m_targa = re.search(
               r'Tipo\s*dato:\s*TARGA\s*[\n\r]+\s*Rif\.\s*testo:\s*(\S+)',
               blocco_text, re.IGNORECASE
           )
           if not m_targa:
               m_targa = re.search(r'(?<!\w)([A-Z]{2}\d{3}[A-Z]{2})(?!\w)', blocco_text, re.IGNORECASE)
           targa_val = m_targa.group(1).strip() if m_targa else ""
           # Prezzo
           m_prezzo = re.search(
               r'[\d,.]+\s+N[12T]\s+([\d]{1,3}(?:[.,]\d{3})*[.,]\d{2})',
               blocco_text, re.IGNORECASE
           )
           prezzo_val = ""
           if m_prezzo:
               prezzo_val = m_prezzo.group(1).replace('.', '').replace(',', '.')
           try:
               if prezzo_val and abs(float(prezzo_val)) == IMPORTO_BOLLO:
                   continue
           except (ValueError, TypeError):
               pass
           if not prezzo_val:
               continue
           print(f"  Riga {match.group(1)}: telaio={telaio_val} targa={targa_val} prezzo={prezzo_val}")
           righe.append({
               "targa":         targa_val,
               "telaio":        telaio_val,
               "descrizione":   desc_val,
               "prezzo_totale": prezzo_val,
           })
   elif is_psa_format:
       row_start_re = re.compile(r'^\s*(\d+)\s*$')
       desc_re = re.compile(r'([A-Z0-9]{8,})\s+(RMK\S+)\s+(.*)', re.IGNORECASE)
       i = 0
       while i < len(lines):
           line = lines[i].strip()
           if row_start_re.match(line):
               block_lines = []
               j = i + 1
               while j < len(lines):
                   next_line = lines[j].strip()
                   if row_start_re.match(next_line) and next_line != line:
                       break
                   block_lines.append(next_line)
                   j += 1
               telaio = ""
               descrizione = ""
               for bl in block_lines:
                   md = desc_re.search(bl)
                   if md:
                       telaio = md.group(1).strip()
                       descrizione = md.group(3).strip()
                       break
               prezzo_totale = ""
               for bl in reversed(block_lines):
                   if re.search(r'\bN\d\b', bl):
                       nums = re.findall(r'[\d]+[.,][\d]+', bl)
                       if nums:
                           prezzo_totale = nums[-1].replace(',', '.')
                       break
               try:
                   if prezzo_totale and abs(float(prezzo_totale)) == IMPORTO_BOLLO:
                       i = j
                       continue
               except (ValueError, TypeError):
                   pass
               if telaio or prezzo_totale:
                   righe.append({"targa": "", "telaio": telaio, "descrizione": descrizione, "prezzo_totale": prezzo_totale})
               i = j
           else:
               i += 1
   elif is_romana_format:
       targa_pattern = re.compile(r'RIF\.TARGA\s+([A-Z0-9]+)', re.IGNORECASE)
       current_targa = ""
       for i, line in enumerate(lines):
           mt = targa_pattern.search(line)
           if mt:
               current_targa = mt.group(1).strip()
               continue
           line_stripped = line.strip()
           if not line_stripped:
               continue
           if re.search(r'bollo', line_stripped, re.IGNORECASE):
               continue
           if re.search(r'imponibile|totale fattura|pagamento|p\.i\.|partita|sede|tel|fax|bonifico|iva|aliq|q\.t[\xc3\xa0a]|prezzo unitario|importo netto|descrizione', line_stripped, re.IGNORECASE):
               continue
           m = re.match(r'^([A-Z][A-Z\s]+?)\s+([\d]{1,3}(?:\.\d{3})*,\d{2})\s*(?:\d+)?$', line_stripped)
           if m and current_targa:
               desc = m.group(1).strip()
               importo = m.group(2).replace('.', '').replace(',', '.')
               try:
                   val = float(importo)
                   if val == IMPORTO_BOLLO:
                       continue
               except ValueError:
                   continue
               righe.append({"targa": current_targa, "telaio": "", "descrizione": desc, "prezzo_totale": importo})
   else:
       for line in lines:
           line_stripped = line.strip()
           if re.search(r'bollo', line_stripped, re.IGNORECASE):
               continue
           m = re.match(r'^(.+?)\s+([\d]{1,3}(?:\.\d{3})*,\d{2})\s*(?:\d+)?$', line_stripped)
           if m:
               desc = m.group(1).strip()
               importo = m.group(2).replace('.', '').replace(',', '.')
               try:
                   val = float(importo)
                   if val == IMPORTO_BOLLO:
                       continue
               except ValueError:
                   continue
               if len(desc) > 2:
                   righe.append({"targa": "", "telaio": "", "descrizione": desc, "prezzo_totale": importo})
   return {"cedente": cedente, "cessionario": cessionario,
           "numero_documento": numero_documento, "data_documento": data_documento, "righe": righe}

def extract_xml_from_pdf(pdf_bytes: bytes):
   """Tenta di estrarre un XML FatturaPA allegato dentro il PDF."""
   try:
       from pypdf import PdfReader
       reader = PdfReader(io.BytesIO(pdf_bytes))
       # Cerca allegati nel PDF
       if '/Names' in reader.trailer['/Root']:
           names = reader.trailer['/Root']['/Names']
           if '/EmbeddedFiles' in names:
               ef = names['/EmbeddedFiles']
               if '/Names' in ef:
                   files = ef['/Names']
                   for i in range(0, len(files), 2):
                       name = str(files[i])
                       filespec = files[i+1].get_object()
                       if '/EF' in filespec:
                           ef_stream = filespec['/EF']['/F'].get_object()
                           data = ef_stream.get_data()
                           if b'FatturaElettronica' in data or b'<?xml' in data:
                               return data
   except Exception as e:
       print(f"Nessun XML allegato trovato nel PDF: {e}")
   return None

def process_pdf_bytes(pdf_bytes, all_rows):
   try:
       # Prima prova a estrarre XML allegato dentro il PDF
       xml_data = extract_xml_from_pdf(pdf_bytes)
       if xml_data:
           print("XML trovato dentro il PDF - uso parser XML")
           return process_xml_bytes(xml_data, all_rows)
       print("Nessun XML allegato - uso parser PDF")

       # Log testo grezzo per debug
       text_preview = ""
       reader = PdfReader(io.BytesIO(pdf_bytes))
       for page in reader.pages:
           t = page.extract_text()
           if t:
               text_preview += t + "\n"
       print("=== PDF TEXT (primi 1500 char) ===")
       print(repr(text_preview[:1500]))
       print("==================================")
       fattura = parse_pdf_fattura(pdf_bytes)
       righe   = fattura.pop("righe", [])
       print(f"Righe estratte: {len(righe)}")
       for r in righe:
           all_rows.append({**fattura, **r})
       return len(righe)
   except Exception as e:
       print(f"Errore PDF: {e}")
       return 0

def process_xml_bytes(xml_bytes, all_rows):
   try:
       fattura = parse_xml_fattura(xml_bytes)
       righe   = fattura.pop("righe", [])
       for r in righe:
           all_rows.append({**fattura, **r})
       return len(righe)
   except Exception as e:
       return 0

def process_zip_bytes(zip_bytes, all_rows, depth=0):
   try:
       with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
           entries = zf.namelist()
           xml_entries = [n for n in entries if n.lower().endswith('.xml')
                         and not n.startswith('__MACOSX')
                         and not n.lower().endswith('signature.xml')]
           pdf_entries = [n for n in entries if n.lower().endswith('.pdf')
                         and not n.startswith('__MACOSX')]
           zip_entries = [n for n in entries if n.lower().endswith('.zip')
                         and not n.startswith('__MACOSX')]
           for xml_name in xml_entries:
               process_xml_bytes(zf.read(xml_name), all_rows)
           for pdf_name in pdf_entries:
               process_pdf_bytes(zf.read(pdf_name), all_rows)
           for zip_name in zip_entries:
               process_zip_bytes(zf.read(zip_name), all_rows, depth + 1)
   except zipfile.BadZipFile:
       pass

# ─── Excel builder ────────────────────────────────────────────────────────────
def valida_targa(targa: str) -> str:
   """Valida e restituisce la targa se valida, altrimenti stringa vuota.
   Formato valido: 2 lettere + 3 cifre + 2 lettere (es. AA000AA), 7 caratteri totali."""
   t = targa.strip().upper()
   if len(t) != 7:
       return ""
   if re.match(r'^[A-Z]{2}\d{3}[A-Z]{2}$', t):
       return t
   return ""

def valida_telaio(telaio: str) -> str:
   """Valida e restituisce il telaio se valido, altrimenti stringa vuota.
   Formato valido: 17 caratteri, inizia con lettere, finisce con numeri."""
   t = telaio.strip().upper()
   if len(t) != 17:
       return ""
   # Inizia con almeno 2 lettere e finisce con almeno 4 cifre
   if re.match(r'^[A-Z]{2,}.*\d{4,}$', t):
       return t
   return ""

def valida_numero_documento(numero: str) -> str:
   """Restituisce il numero documento se valido (numerico o alfanumerico tipo G000617)."""
   n = numero.strip()
   # Accetta numeri puri (000866019) o alfanumerici (G000617)
   # Esclude parole come "DOCUMENTO", "ART", ecc.
   if re.match(r'^[A-Z0-9]{4,}$', n, re.IGNORECASE) and not re.match(r'^[A-Z]+$', n, re.IGNORECASE):
       return n
   return ""

def to_float(val):
   try:
       return float(str(val).replace(',', '.'))
   except (ValueError, AttributeError):
       return None

def get_categoria(descrizione: str) -> str:
   """Estrae la categoria dalla descrizione, solo parole chiave specifiche."""
   d = descrizione.lower()
   if "forfait" in d:
       return "Forfait"
   if "over plafond" in d:
       return "Over Plafond"
   if "km eccedenti" in d or "esubero km" in d:
       return "Km Eccedenti"
   if "eam" in d:
       return "EAM"
   if "tagliando" in d:
       return "Tagliando"
   if "perizia" in d:
       return "Perizia"
   return ""

def build_row(row):
   descrizione = row.get("descrizione", "")
   targa  = valida_targa(row.get("targa", ""))
   telaio = valida_telaio(row.get("telaio", ""))
   return [
       row.get("cedente", ""),
       row.get("cessionario", ""),
       valida_numero_documento(row.get("numero_documento", "")),
       row.get("data_documento", ""),
       targa,
       telaio,
       to_float(row.get("prezzo_totale", "")),
       descrizione,
       get_categoria(descrizione),
   ]

def write_header(ws, headers, col_widths):
   header_fill = PatternFill("solid", fgColor=HEADER_BG)
   header_font = Font(bold=True, color=HEADER_FONT)
   for c, (h, w) in enumerate(zip(headers, col_widths), start=1):
       cell = ws.cell(row=1, column=c, value=h)
       cell.fill = header_fill
       cell.font = header_font
       cell.alignment = Alignment(horizontal="center", vertical="center")
       ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
   ws.row_dimensions[1].height = 20
   ws.freeze_panes = "A2"

def build_excel(all_rows):
   headers    = ["Cedente", "Cessionario", "N. Documento", "Data Documento",
                 "Targa", "Telaio", "Prezzo Totale (€)",
                 "Descrizione (Forfait/Over Plafond/ETM/EAM/KM…)",
                 "Descrizione"]
   col_widths = [30, 30, 18, 16, 14, 22, 18, 45, 18]
   CATEGORIE = ["EAM", "Forfait", "Km Eccedenti", "Over Plafond", "Tagliando"]
   # Filtra BOLLO
   filtered = [r for r in all_rows
               if "bollo" not in r.get("descrizione", "").lower()
               and "bollo" not in r.get("telaio", "").lower()]
   # Duplicati
   seen, uniq, dups = {}, [], []
   for r in filtered:
       key = (r.get("cedente",""), r.get("cessionario",""),
              r.get("numero_documento",""), r.get("data_documento",""),
              r.get("telaio",""), r.get("prezzo_totale",""), r.get("descrizione",""))
       if key in seen:
           dups.append(r)
       else:
           seen[key] = True
           uniq.append(r)
   wb = openpyxl.Workbook()
   # ── Foglio 1: Fatture ──
   ws1 = wb.active
   ws1.title = "Fatture"
   write_header(ws1, headers, col_widths)
   for row in uniq:
       ws1.append(build_row(row))
   # ── Foglio 2: Duplicati ──
   ws2 = wb.create_sheet(title="Duplicati")
   write_header(ws2, headers, col_widths)
   for row in dups:
       ws2.append(build_row(row))
   out = io.BytesIO()
   wb.save(out)
   out.seek(0)
   return out, len(uniq), len(dups)

# ─── HTML Template ────────────────────────────────────────────────────────────
HTML = '''<!DOCTYPE html>
<html lang="it">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Estrattore Fatture — Hertz</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=Space+Grotesk:wght@500;600;700&display=swap" rel="stylesheet"/>
<style>
*{box-sizing:border-box;margin:0;padding:0}
:root{
 --bg:#0a0a0f;
 --surface:#111118;
 --surface2:#1a1a24;
 --border:#ffffff0f;
 --border2:#ffffff18;
 --accent:#ff6b35;
 --accent2:#ff9a6c;
 --green:#00e5a0;
 --red:#ff4d6d;
 --text:#f0f0f5;
 --muted:#7070a0;
 --dim:#3a3a55;
}
body{
 background:var(--bg);
 color:var(--text);
 font-family:'Inter',sans-serif;
 min-height:100vh;
 display:flex;
 flex-direction:column;
 align-items:center;
 padding:0 0 80px;
}
::-webkit-scrollbar{width:4px}
::-webkit-scrollbar-thumb{background:var(--accent);border-radius:2px}
/* HERO */
.hero{
 width:100%;
 background:linear-gradient(135deg,#0a0a0f 0%,#12101e 50%,#0f1218 100%);
 border-bottom:1px solid var(--border2);
 padding:60px 24px 48px;
 text-align:center;
 position:relative;
 overflow:hidden;
}
.hero::before{
 content:'';
 position:absolute;
 top:-120px;left:50%;transform:translateX(-50%);
 width:600px;height:300px;
 background:radial-gradient(ellipse,rgba(255,107,53,0.15) 0%,transparent 70%);
 pointer-events:none;
}
.hero-badge{
 display:inline-flex;align-items:center;gap:8px;
 background:rgba(255,107,53,0.1);
 border:1px solid rgba(255,107,53,0.3);
 color:var(--accent2);
 font-size:12px;font-weight:600;letter-spacing:1.5px;text-transform:uppercase;
 padding:6px 16px;border-radius:100px;
 margin-bottom:24px;
}
.hero-badge::before{content:'●';font-size:8px;color:var(--accent)}
h1{
 font-family:'Space Grotesk',sans-serif;
 font-size:clamp(32px,5vw,56px);
 font-weight:700;
 color:#fff;
 letter-spacing:-1.5px;
 line-height:1.1;
 margin-bottom:12px;
}
h1 span{
 background:linear-gradient(90deg,var(--accent),var(--accent2));
 -webkit-background-clip:text;-webkit-text-fill-color:transparent;
}
.hero-sub{
 font-size:16px;color:var(--muted);font-weight:400;
 max-width:480px;margin:0 auto;line-height:1.6;
}
.hero-stats{
 display:flex;gap:32px;justify-content:center;margin-top:40px;flex-wrap:wrap;
}
.stat{text-align:center}
.stat-val{font-family:'Space Grotesk',sans-serif;font-size:24px;font-weight:700;color:#fff}
.stat-label{font-size:12px;color:var(--muted);margin-top:2px}
/* MAIN */
.main{width:100%;max-width:680px;padding:40px 24px 0;display:flex;flex-direction:column;gap:20px}
/* CARD */
.card{
 background:var(--surface);
 border:1px solid var(--border);
 border-radius:16px;
 overflow:hidden;
}
.card-header{
 padding:16px 20px;
 border-bottom:1px solid var(--border);
 display:flex;align-items:center;gap:10px;
}
.card-icon{
 width:32px;height:32px;border-radius:8px;
 background:rgba(255,107,53,0.1);
 display:flex;align-items:center;justify-content:center;
 font-size:16px;
}
.card-title{font-size:14px;font-weight:600;color:#fff}
.card-sub{font-size:12px;color:var(--muted);margin-top:1px}
/* UPLOAD */
.drop-zone{
 padding:32px 24px;
 display:flex;flex-direction:column;align-items:center;gap:10px;
 cursor:pointer;
 transition:.2s;
 border-bottom:1px solid var(--border);
 background:transparent;
}
.drop-zone:hover,.drop-zone.over{background:rgba(255,107,53,0.04)}
.drop-icon-wrap{
 width:56px;height:56px;border-radius:16px;
 background:linear-gradient(135deg,rgba(255,107,53,0.15),rgba(255,107,53,0.05));
 border:1px solid rgba(255,107,53,0.2);
 display:flex;align-items:center;justify-content:center;
 font-size:24px;
 transition:.2s;
}
.drop-zone:hover .drop-icon-wrap{
 background:linear-gradient(135deg,rgba(255,107,53,0.25),rgba(255,107,53,0.1));
 transform:translateY(-2px);
}
.drop-text{font-size:14px;font-weight:500;color:var(--text)}
.drop-sub{font-size:12px;color:var(--muted)}
.or-row{
 display:flex;align-items:center;gap:12px;
 padding:0 24px;
 font-size:11px;color:var(--dim);letter-spacing:1px;text-transform:uppercase;
}
.or-row::before,.or-row::after{content:'';flex:1;height:1px;background:var(--border)}
.folder-btn{
 padding:16px 24px;
 display:flex;align-items:center;gap:12px;
 cursor:pointer;transition:.2s;
 background:transparent;font-family:'Inter',sans-serif;
 color:var(--muted);font-size:13px;width:100%;
}
.folder-btn:hover{background:rgba(255,255,255,0.03);color:var(--text)}
.folder-icon{
 width:36px;height:36px;border-radius:10px;
 background:var(--surface2);border:1px solid var(--border);
 display:flex;align-items:center;justify-content:center;font-size:18px;
 flex-shrink:0;
}
.folder-text{text-align:left}
.folder-label{font-weight:500;color:var(--text);font-size:13px}
.folder-desc{font-size:11px;color:var(--muted);margin-top:2px}
/* FILE LIST */
.file-panel{background:var(--surface);border:1px solid var(--border);border-radius:16px;overflow:hidden;display:none}
.file-panel.show{display:block}
.file-panel-header{
 padding:12px 16px;border-bottom:1px solid var(--border);
 display:flex;justify-content:space-between;align-items:center;
 background:var(--surface2);
}
.file-count{font-size:12px;font-weight:600;color:var(--accent)}
.clear-btn{
 background:transparent;border:1px solid var(--border);color:var(--muted);
 font-size:11px;padding:4px 10px;border-radius:6px;cursor:pointer;font-family:'Inter',sans-serif;
 transition:.15s;
}
.clear-btn:hover{color:var(--red);border-color:var(--red)}
.file-list{max-height:160px;overflow-y:auto;padding:4px 0}
.file-row{
 display:flex;justify-content:space-between;align-items:center;
 padding:8px 16px;transition:.15s;
}
.file-row:hover{background:rgba(255,255,255,0.02)}
.file-name{font-size:12px;color:var(--muted);overflow:hidden;text-overflow:ellipsis;white-space:nowrap;flex:1}
.file-name::before{content:'📄 '}
.rm-btn{background:transparent;border:none;color:var(--dim);cursor:pointer;font-size:13px;padding:2px 6px;border-radius:4px;transition:.15s}
.rm-btn:hover{color:var(--red);background:rgba(255,77,109,0.1)}
/* RUN BUTTON */
.run-btn{
 width:100%;
 background:linear-gradient(135deg,var(--accent),#ff8c5a);
 color:#fff;
 font-family:'Space Grotesk',sans-serif;font-weight:600;font-size:15px;
 border:none;border-radius:12px;padding:16px;
 cursor:pointer;transition:.2s;
 box-shadow:0 4px 20px rgba(255,107,53,0.3);
 position:relative;overflow:hidden;
}
.run-btn::before{
 content:'';position:absolute;top:0;left:-100%;width:100%;height:100%;
 background:linear-gradient(90deg,transparent,rgba(255,255,255,0.1),transparent);
 transition:.5s;
}
.run-btn:hover:not(:disabled)::before{left:100%}
.run-btn:hover:not(:disabled){transform:translateY(-1px);box-shadow:0 8px 30px rgba(255,107,53,0.4)}
.run-btn:disabled{opacity:0.4;cursor:not-allowed;transform:none}
/* PROGRESS */
.progress-card{background:var(--surface);border:1px solid var(--border);border-radius:16px;padding:20px;display:none}
.progress-card.show{display:block}
.progress-label{font-size:12px;color:var(--muted);margin-bottom:10px;display:flex;justify-content:space-between}
.progress-track{width:100%;height:4px;background:var(--surface2);border-radius:2px;overflow:hidden}
.progress-bar{height:100%;background:linear-gradient(90deg,var(--accent),var(--accent2));border-radius:2px;width:0%;transition:width .4s ease}
/* LOG */
.log-card{background:var(--surface);border:1px solid var(--border);border-radius:16px;overflow:hidden;display:none}
.log-card.show{display:block}
.log-header{padding:12px 16px;border-bottom:1px solid var(--border);background:var(--surface2);font-size:12px;font-weight:600;color:var(--muted);letter-spacing:.5px;text-transform:uppercase}
.log-body{padding:16px;max-height:220px;overflow-y:auto;display:flex;flex-direction:column;gap:4px}
.log-line{font-size:12px;font-family:'SF Mono','Fira Code',monospace;line-height:1.7;display:flex;gap:8px;align-items:baseline}
.log-line::before{content:'›';color:var(--dim);flex-shrink:0}
.log-info{color:#8888bb}
.log-ok{color:var(--green)}
.log-err{color:var(--red)}
.log-zip{color:var(--accent)}
/* DONE */
.done-card{
 background:linear-gradient(135deg,rgba(0,229,160,0.08),rgba(0,229,160,0.03));
 border:1px solid rgba(0,229,160,0.2);
 border-radius:16px;padding:24px;
 display:none;text-align:center;
}
.done-card.show{display:block}
.done-icon{font-size:36px;margin-bottom:12px}
.done-title{font-family:'Space Grotesk',sans-serif;font-size:18px;font-weight:600;color:var(--green);margin-bottom:6px}
.done-sub{font-size:13px;color:var(--muted)}
.done-stats{display:flex;gap:20px;justify-content:center;margin-top:16px;flex-wrap:wrap}
.done-stat{
 background:rgba(0,229,160,0.08);border:1px solid rgba(0,229,160,0.15);
 border-radius:8px;padding:10px 20px;text-align:center;
}
.done-stat-val{font-family:'Space Grotesk',sans-serif;font-size:20px;font-weight:700;color:var(--green)}
.done-stat-label{font-size:11px;color:var(--muted);margin-top:2px}
/* FOOTER */
.footer{
 margin-top:40px;text-align:center;font-size:11px;color:var(--dim);
 padding:0 24px;
}
.footer a{color:var(--muted);text-decoration:none}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.4}}
.pulsing{animation:pulse 1.2s infinite}
@keyframes fadeIn{from{opacity:0;transform:translateY(6px)}to{opacity:1;transform:translateY(0)}}
.fade-in{animation:fadeIn .3s ease}
</style>
</head>
<body>
<!-- HERO -->
<div class="hero">
<div class="hero-badge">Estrattore Fatture</div>
<h1>Da fattura a <span>Excel</span><br>in pochi secondi</h1>
<p class="hero-sub">Carica XML, PDF o ZIP — il sistema estrae automaticamente tutti i dati strutturati.</p>
<div class="hero-stats">
<div class="stat"><div class="stat-val">3</div><div class="stat-label">Formati supportati</div></div>
<div class="stat"><div class="stat-val">∞</div><div class="stat-label">File per volta</div></div>
<div class="stat"><div class="stat-val">2</div><div class="stat-label">Fogli Excel</div></div>
</div>
</div>
<!-- MAIN -->
<div class="main">
<!-- UPLOAD CARD -->
<div class="card">
<div class="card-header">
<div class="card-icon">📂</div>
<div>
<div class="card-title">Carica i tuoi file</div>
<div class="card-sub">XML, PDF o ZIP · anche più file insieme</div>
</div>
</div>
<label class="drop-zone" id="dropZone">
<input type="file" id="fileInput" accept=".xml,.zip,.pdf" multiple style="display:none"/>
<div class="drop-icon-wrap">⬇</div>
<div class="drop-text">Trascina qui i file oppure clicca</div>
<div class="drop-sub">XML · PDF · ZIP supportati</div>
</label>
<div class="or-row">oppure</div>
<label class="folder-btn">
<input type="file" id="folderInput" style="display:none" webkitdirectory directory/>
<div class="folder-icon">🗂</div>
<div class="folder-text">
<div class="folder-label">Carica una cartella intera</div>
<div class="folder-desc">Prende automaticamente tutti gli XML e ZIP al suo interno</div>
</div>
</label>
</div>
<!-- FILE LIST -->
<div class="file-panel" id="filePanel">
<div class="file-panel-header">
<span class="file-count" id="fpCount">0 file selezionati</span>
<button class="clear-btn" onclick="clearAll()">✕ Svuota tutto</button>
</div>
<div class="file-list" id="fileList"></div>
</div>
<!-- RUN -->
<button class="run-btn" id="runBtn" onclick="handleRun()" disabled>
   ▶ &nbsp;Avvia Estrazione
</button>
<!-- PROGRESS -->
<div class="progress-card" id="progressWrap">
<div class="progress-label">
<span>Elaborazione in corso...</span>
<span id="progressPct">0%</span>
</div>
<div class="progress-track"><div class="progress-bar" id="progressBar"></div></div>
</div>
<!-- LOG -->
<div class="log-card" id="logCard">
<div class="log-header">Log di sistema</div>
<div class="log-body" id="logBox"></div>
</div>
<!-- DONE -->
<div class="done-card" id="doneBanner">
<div class="done-icon">✅</div>
<div class="done-title">Estrazione completata!</div>
<div class="done-sub">Il file <strong>estrazione_fatture.xlsx</strong> è stato scaricato</div>
<div class="done-stats" id="doneStats"></div>
</div>
<div class="footer">
   I file vengono elaborati sul server e non vengono salvati. &nbsp;·&nbsp;
<a href="#">estrattore-fatture.onrender.com</a>
</div>
</div>
<script>
let selectedFiles = [];
function addFiles(newFiles) {
 const valid = Array.from(newFiles).filter(f =>
   ['xml','zip','pdf'].some(ext => f.name.toLowerCase().endsWith('.'+ext))
 );
 const existing = new Set(selectedFiles.map(f => f.name+f.size));
 valid.forEach(f => { if(!existing.has(f.name+f.size)) selectedFiles.push(f); });
 render();
}
function removeFile(i){ selectedFiles.splice(i,1); render(); }
function clearAll(){
 selectedFiles=[];render();
 document.getElementById('logCard').classList.remove('show');
 document.getElementById('doneBanner').classList.remove('show');
}
function render(){
 const panel=document.getElementById('filePanel');
 const list=document.getElementById('fileList');
 const count=document.getElementById('fpCount');
 const btn=document.getElementById('runBtn');
 if(!selectedFiles.length){panel.classList.remove('show');btn.disabled=true;return;}
 panel.classList.add('show');btn.disabled=false;
 count.textContent=selectedFiles.length+' file selezionati';
 list.innerHTML=selectedFiles.map((f,i)=>`
<div class="file-row fade-in">
<span class="file-name">${f.name}</span>
<button class="rm-btn" onclick="removeFile(${i})">✕</button>
</div>`).join('');
}
document.getElementById('fileInput').onchange=e=>addFiles(e.target.files);
document.getElementById('folderInput').onchange=e=>addFiles(e.target.files);
const dz=document.getElementById('dropZone');
dz.addEventListener('dragover',e=>{e.preventDefault();dz.classList.add('over');});
dz.addEventListener('dragleave',()=>dz.classList.remove('over'));
dz.addEventListener('drop',e=>{e.preventDefault();dz.classList.remove('over');addFiles(e.dataTransfer.files);});
dz.addEventListener('click',()=>document.getElementById('fileInput').click());
function log(msg,type='info'){
 const box=document.getElementById('logBox');
 document.getElementById('logCard').classList.add('show');
 const d=document.createElement('div');
 d.className='log-line log-'+type+' fade-in';
 d.textContent=msg;
 box.appendChild(d);
 box.scrollTop=box.scrollHeight;
}
function setProgress(pct){
 document.getElementById('progressBar').style.width=pct+'%';
 document.getElementById('progressPct').textContent=pct+'%';
}
async function handleRun(){
 if(!selectedFiles.length) return;
 const btn=document.getElementById('runBtn');
 btn.disabled=true;btn.classList.add('pulsing');
 btn.innerHTML='⏳ &nbsp;Elaborazione in corso…';
 document.getElementById('logBox').innerHTML='';
 document.getElementById('logCard').classList.add('show');
 document.getElementById('doneBanner').classList.remove('show');
 document.getElementById('progressWrap').classList.add('show');
 setProgress(20);
 const fd=new FormData();
 selectedFiles.forEach(f=>fd.append('files',f));
 try{
   log('Invio '+selectedFiles.length+' file al server…','info');
   const resp=await fetch('/process',{method:'POST',body:fd});
   setProgress(85);
   if(!resp.ok){
     const err=await resp.json();
     log('Errore: '+(err.error||resp.statusText),'err');
     return;
   }
   const fatture=resp.headers.get('X-Rows-Fatture')||'?';
   const dups=resp.headers.get('X-Rows-Duplicati')||'?';
   log('Elaborazione completata con successo','ok');
   log('Foglio Fatture: '+fatture+' righe','ok');
   log('Foglio Duplicati: '+dups+' righe','ok');
   setProgress(100);
   const blob=await resp.blob();
   const url=URL.createObjectURL(blob);
   const a=document.createElement('a');
   a.href=url;a.download='estrazione_fatture.xlsx';
   document.body.appendChild(a);a.click();
   document.body.removeChild(a);URL.revokeObjectURL(url);
   const banner=document.getElementById('doneBanner');
   document.getElementById('doneStats').innerHTML=`
<div class="done-stat"><div class="done-stat-val">${fatture}</div><div class="done-stat-label">Righe fatture</div></div>
<div class="done-stat"><div class="done-stat-val">${dups}</div><div class="done-stat-label">Duplicati trovati</div></div>`;
   banner.classList.add('show','fade-in');
 }catch(e){
   log('Errore di rete: '+e.message,'err');
 }finally{
   btn.disabled=false;btn.classList.remove('pulsing');
   btn.innerHTML='▶ &nbsp;Avvia Estrazione';
 }
}
</script>
</body>
</html>'''

@app.route('/debug', methods=['POST'])
def debug():
   """Endpoint di debug: restituisce il testo grezzo estratto dal PDF."""
   files = request.files.getlist('files')
   if not files:
       return jsonify({"error": "Nessun file"}), 400
   f = files[0]
   data = f.read()
   text = ""
   try:
       with pdfplumber.open(io.BytesIO(data)) as pdf:
           for page in pdf.pages:
               t = page.extract_text()
               if t:
                   text += t + "\n"
   except Exception as e:
       return jsonify({"error": str(e)}), 500
   return jsonify({"text": text[:3000]})

@app.route('/')
def index():
   return render_template_string(HTML)

@app.route('/process', methods=['POST'])
def process():
   files = request.files.getlist('files')
   if not files:
       return jsonify({"error": "Nessun file ricevuto"}), 400
   all_rows = []
   for f in files:
       name = f.filename.lower()
       data = f.read()
       if name.endswith('.xml'):
           process_xml_bytes(data, all_rows)
       elif name.endswith('.pdf'):
           process_pdf_bytes(data, all_rows)
       elif name.endswith('.zip'):
           process_zip_bytes(data, all_rows)
   if not all_rows:
       return jsonify({"error": "Nessun dato estratto dai file caricati"}), 422
   excel_bytes, n_fatture, n_dups = build_excel(all_rows)
   response = send_file(
       excel_bytes,
       mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
       as_attachment=True,
       download_name='estrazione_fatture.xlsx'
   )
   response.headers['X-Rows-Fatture']   = str(n_fatture)
   response.headers['X-Rows-Duplicati'] = str(n_dups)
   response.headers['Access-Control-Expose-Headers'] = 'X-Rows-Fatture, X-Rows-Duplicati'
   return response

if __name__ == '__main__':
   port = int(os.environ.get('PORT', 5000))
   app.run(host='0.0.0.0', port=port)
